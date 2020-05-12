from vk_api.longpoll import VkLongPoll, VkEventType
import vk_api
from openpyxl import Workbook
from openpyxl import load_workbook
from base import affirmative
from base import start
from base import stop
from datetime import datetime
import time
def alpabet(x): return "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[x]


token = "374eeed4f9510e8e6c2e5fbfbaab5f93c8068af27a245c2f729583018f34d608e7d740e2d349cf2d28997"
vk_session = vk_api.VkApi(token=token)
longpoll = VkLongPoll(vk_session)
index = 0
list = []
stage = 1

def handle_first_message(event, ws):
    try:
        load_workbook(str(event.user_id) + ".xlsx")
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "С возвращением!",
                                            'random_id': 0})
        return 5
    except:
        print("no")
    response = event.text.lower()
    if response in start:
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Добро пожаловать!",
                                            'random_id': 0})
        time.sleep(1)
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Введите свои дисциплины:",
                                            'random_id': 0})
        ws['A1'].value = event.user_id
        return 2
    return 1

def list_disciplines(list):
    disciplines = ""
    for i in range(0, len(list)):
        disciplines += str(list[i])
        disciplines += "\n"
    return disciplines


def get_discipline(discipline, list):
    if discipline.lower() in stop:
        return 0
    list.append(str(discipline))
    return list

def handle_disciplines(event, list):
    if get_discipline(event.text, list) == 0:
        if len(list) == 0:
            vk_session.method("messages.send", {'user_id': event.user_id,
                                                'message': "Вы не ввели ни одной дисциплины!",
                                                'random_id': 0})
            time.sleep(1)
            vk_session.method("messages.send", {'user_id': event.user_id,
                                                'message': "Введите свои дисциплины:",
                                                'random_id': 0})
            return 2
        disc = list_disciplines(list)
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Я все правильно записал?\n" + disc,
                                            'random_id': 0})
        return 3
    return 2

def check_answer(event, list):
    if event.text.lower() in affirmative:
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Отлично, продолжим",
                                            'random_id': 0})
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Введите желаемый балл для дисциплины " + str(list[0]),
                                            'random_id': 0})
        return 4
    else:
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Начнем заново",
                                            'random_id': 0})
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Введите свои дисциплины:",
                                            'random_id': 0})
        return 2

def fill_in_the_goals(event, list, index, ws):
    try:
        int(event.text)
    except:
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Введите числовое значение для предмета " + str(list[index]),
                                            'random_id': 0})
        return index
    letter = alpabet(index + 1)
    ws[letter + '2'].value = list[index]
    ws[letter + '3'].value = int(event.text)
    ws[letter + '4'].value = int(0)
    index += 1
    if len(list) - index > 0:
        vk_session.method("messages.send", {'user_id': event.user_id,
                                        'message': "Введите желаемый балл для дисциплины " + str(list[index]),
                                        'random_id': 0})
    return index

def make_string(ws, flag):
    i = 1
    string = "\n"
    while ws[alpabet(i) + "2"].value:
        if flag == 0:
            string += str(ws[alpabet(i) + "2"].value) + " - " + str(ws[alpabet(i) + "3"].value) + "\n"
        elif flag == 1:
            string += str(ws[alpabet(i) + "2"].value) + " "
        # elif flag == 2:
        #     string += str(ws[alpabet(i) + "2"].value) + " - " + str(ws[alpabet(i) + "3"].value) + "\n"
        i += 1
    if flag == 0:
        string = string[0:-1]
        string += "?"
    return string


def check_point(ws):
    string = make_string(ws, 0)
    vk_session.method("messages.send", {'user_id': event.user_id,
                                        'message': "Проверим результаты:" + string,
                                        'random_id': 0})


def analize_query(words_in_message, ws):
    i = 1
    letter = alpabet(1)
    print("здеся\n")
    print(words_in_message)
    while ws[letter + "2"].value and i < 26:
        if ws[letter + "2"].value == words_in_message[0]:
            break
        i += 1
        letter = alpabet(i)
    if len(words_in_message) > 1:
        ws[letter + "4"].value += int(words_in_message[1])
    return str(ws[letter + "2"].value) + " " + str(ws[letter + "4"].value) + "/" + str(ws[letter + "3"].value)

def print_all(string, ws):
    string = string.split()
    i = 0
    output = ""
    while i < len(string):
        output += analize_query([string[i]], ws) + "\n"
        i += 1
    return output

def write_or_unload(event):
    wb = load_workbook(str(event.user_id) + ".xlsx")
    ws = wb.active
    words_in_message = event.text.split()
    list = make_string(ws, 1)
    print(list)
    if words_in_message[0] in list:
        string = analize_query(words_in_message, ws)
        wb.save(str(ws['A1'].value) + ".xlsx")
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': string,
                                            'random_id': 0})
    elif event.text in stop:
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Пока",
                                            'random_id': 0})
        wb.save(str(ws['A1'].value) + ".xlsx")
        return 0
    elif event.text == "Все":
        string = print_all(make_string(ws, 1), ws)
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': string,
                                            'random_id': 0})
    else:
        vk_session.method("messages.send", {'user_id': event.user_id,
                                            'message': "Дисциплина " + words_in_message[0] + " не найдена",
                                            'random_id': 0})

wb = Workbook()
ws = wb.active

x = 1


while x == 1:
    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW:
            print("Сообщение пришло в: " + str(datetime.strftime(datetime.now(), "%H:%M:%S")))
            print("Текст сообщения: " + str(event.text))
            if not event.from_me:
                if stage == 5:
                    x = write_or_unload(event)
                    if x == 0:
                        break
                if stage == 4 and index <= len(list):
                    if index < len(list):
                        index = fill_in_the_goals(event, list, index, ws) # записывает в соответствие каждому предмету цель по нему
                    if index == len(list):
                        if event.text in affirmative:
                            wb.save(str(ws['A1'].value) + ".xlsx")
                            vk_session.method("messages.send", {'user_id': event.user_id,
                                                                'message': "Данные были занесены",
                                                                'random_id': 0})
                            stage = 5
                        elif event.text == "нет":
                            index = 0
                            vk_session.method("messages.send", {'user_id': event.user_id,
                                                                'message': "Введите желаемый балл для дисциплины " + str(list[index]),
                                                                'random_id': 0})
                        else:
                            check_point(ws)
                if stage == 3:
                    stage = check_answer(event, list) # проверяет правильно ли чел ввел свои предметы, предлагает перезабить
                    if stage == 2:                      # если перезабивает, то удаляет все, что записал ранее и возвращает на стадию 2
                        list = []
                        continue
                if stage == 2:
                    stage = handle_disciplines(event, list) # заносит все дисциплины в list
                if stage == 1:
                    stage = handle_first_message(event, ws) # первое сообщение, запуск бота